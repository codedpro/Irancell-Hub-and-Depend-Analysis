import openpyxl
import re

def clean(filename, output_filename):
    wb = openpyxl.load_workbook(filename)
    cleaned_wb = openpyxl.Workbook()

    pattern = re.compile(r'[a-zA-Z]\d{4}')

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        cleaned_sheet = cleaned_wb.create_sheet(title=sheet_name)

        for row in sheet.iter_rows():
            cleaned_row = []
            for cell in row:
                cell_value = str(cell.value)
                matches = list(pattern.finditer(cell_value))
                
                if matches:
                    if len(matches) == 1:
                        cleaned_value = matches[0].group()
                    else:
                        first_match = matches[0].group()
                        second_match = matches[1].group()
                        
                        if first_match.startswith('P'):
                            cleaned_value = second_match
                        else:
                            cleaned_value = first_match
                    cleaned_row.append(cleaned_value)
            cleaned_sheet.append(cleaned_row)

    default_sheet = cleaned_wb.active
    cleaned_wb.remove(default_sheet)

    cleaned_wb.save(output_filename)

